"""Bulk sheet importer (V2 R2-2) — B2B client rosters and historical backfill.

Design choice worth explaining: dry_run() runs the *real* row-processing path
(resolve_or_create_contact + register, the exact same functions the public
endpoint and the desk use) inside a SAVEPOINT that gets rolled back at the
end, rather than a separate "simulate" implementation. This is what makes
"dry-run counts == commit counts" a structural guarantee instead of a hope —
there's only one code path, dry-run just throws its writes away afterward.
It also correctly handles two rows in the same sheet sharing an email/phone
(a duplicate the operator didn't notice): row 2 sees row 1's contact, which
was flushed (not committed) earlier in the same nested transaction, exactly
as it would at real commit time.

commit() never re-parses the uploaded file — it replays the exact parsed,
evaluated row data dry_run() already stored on the batch. The file itself
isn't kept; only its extracted fields are, in import_batches.counts (JSONB).
"""

from __future__ import annotations

import io
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from typing import Literal
from uuid import UUID, uuid4

import openpyxl
from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.sessions.cohort import Cohort
from app.models.sessions.import_batch import ImportBatch
from app.models.sessions.registration import Registration
from app.models.spine.contact import ContactRelationship
from app.services.spine.identity import ARABIC_INDIC_DIGITS, normalize_phone, resolve_or_create_contact
from app.services.sessions.registration import register

# (column_name, required) — order matches the downloadable template. No
# separate Arabic-name column — a name is just a name. date_of_birth/grade/
# organization (2026-07-24, CEO request) are purely informational, optional,
# and not enforced against anywhere in this system.
TEMPLATE_COLUMNS: list[tuple[str, bool]] = [
    ("student_name", True),
    ("email", True),
    ("phone", True),
    ("city", True),
    ("date_of_birth", False),
    ("grade", False),
    ("organization", False),
    ("parent_name", False),
    ("parent_phone", False),
    ("parent_email", False),
]
_REQUIRED_COLUMNS = [name for name, required in TEMPLATE_COLUMNS if required]


class _DryRunRollback(Exception):
    """Sentinel to force the dry-run's SAVEPOINT to roll back once every row
    has been processed and recorded — never meant to escape dry_run()."""


@dataclass
class RowResult:
    row_number: int
    disposition: Literal["create", "link", "already_registered", "review", "error"]
    data: dict
    reason: str | None = None
    contact_id: str | None = None  # only set for link/already_registered (stable across the rollback)


def _normalize_header(raw: str) -> str:
    return str(raw or "").strip().lower().replace(" ", "_").replace("-", "_")


def _find_header_row(sheet) -> tuple[int, dict[str, int]]:
    """Scan the first 5 rows for one whose normalized cells cover every
    required column — tolerates a title row, a blank row, or a logo/banner
    row above the real header."""
    expected = {name for name, _ in TEMPLATE_COLUMNS}
    for row_idx in range(1, 6):
        cells = [_normalize_header(c.value) for c in sheet[row_idx]]
        col_index = {name: i for i, name in enumerate(cells) if name in expected}
        if all(name in col_index for name in _REQUIRED_COLUMNS):
            return row_idx, col_index
    raise HTTPException(
        status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail=f"Couldn't find a header row with all required columns ({', '.join(_REQUIRED_COLUMNS)}) in the first 5 rows",
    )


def parse_workbook(file_bytes: bytes) -> list[dict]:
    """Returns one dict per data row, keys = canonical column names, values =
    raw cell values (still unvalidated strings/dates/numbers)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.worksheets[0]
    header_row, col_index = _find_header_row(sheet)

    rows = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        cells = sheet[row_idx]
        if all(c.value is None for c in cells):
            continue  # skip fully blank rows
        row = {name: cells[idx].value for name, idx in col_index.items()}
        rows.append({"row_number": row_idx, "raw": row})
    return rows


def _clean_str(value) -> str | None:
    """Cell value -> trimmed string, or None if empty.

    Excel stores anything that looks numeric as a number, so a phone typed as
    `1119394400` comes back from openpyxl as the float 1119394400.0 and a naive
    str() yields "1119394400.0" — which no phone parser will accept, and which
    produced the confusing "could not be parsed as a valid number" error on
    perfectly reasonable-looking sheets. Whole-number floats are narrowed back
    to int before stringifying. (The real fix for users is formatting the
    column as Text, but the importer shouldn't punish them for not knowing.)
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).translate(ARABIC_INDIC_DIGITS).strip()
    return s or None


def _parse_date_of_birth(value) -> date | None:
    """Lenient, best-effort — date_of_birth is optional and purely
    informational (2026-07-24), so an unparseable value is silently dropped
    rather than failing the whole row. By the time this runs, a real Excel
    date cell has already gone through _json_safe_raw and arrived as an ISO
    string, not a date/datetime object.

    A cell the author never formatted as a date arrives as an Excel serial
    number instead (days since 1899-12-30), which none of the string parsers
    below can read — so that's handled explicitly. The range guard keeps it
    from turning an unrelated number into a spurious birthday: 1000-80000 is
    roughly 1902-2119.
    """
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 1000 <= float(value) <= 80000:
            return date(1899, 12, 30) + timedelta(days=int(value))
        return None

    s = _clean_str(value)
    if not s:
        return None
    for parser in (date.fromisoformat, lambda v: datetime.strptime(v, "%d/%m/%Y").date(),
                   lambda v: datetime.strptime(v, "%m/%d/%Y").date(),
                   lambda v: datetime.strptime(v, "%Y/%m/%d").date(),
                   lambda v: datetime.strptime(v, "%d-%m-%Y").date()):
        try:
            return parser(s)
        except ValueError:
            continue
    return None


async def _ensure_guardian_relationship(db: AsyncSession, *, student_id: UUID, guardian_id: UUID) -> None:
    result = await db.execute(
        select(ContactRelationship).where(
            ContactRelationship.contact_id == guardian_id,
            ContactRelationship.related_contact_id == student_id,
            ContactRelationship.relation == "guardian_of",
        )
    )
    if result.scalars().first() is not None:
        return
    db.add(ContactRelationship(id=uuid4(), contact_id=guardian_id, related_contact_id=student_id, relation="guardian_of"))
    await db.flush()


def _json_safe_raw(raw: dict) -> dict:
    """The raw row as pulled from openpyxl, made JSONB-storable — date/datetime
    cells become ISO strings. This is the ONE row shape used everywhere:
    parsed fresh from the sheet at dry-run time, and replayed byte-for-byte
    (via this same JSON-safe form, stored on the batch) at commit time. There
    is no separate "already parsed" shape — that was a bug caught before it
    shipped: storing the parsed output instead of the raw input meant
    commit() was calling _process_row on a dict missing keys it expects,
    silently turning every successful dry-run row into a commit-time error.
    """
    safe = {}
    for k, v in raw.items():
        if isinstance(v, (date, datetime)):
            safe[k] = v.isoformat()
        else:
            safe[k] = v
    return safe


async def _process_row(
    db: AsyncSession, row_number: int, raw: dict, cohort: Cohort, payment_status: str, set_contact_organization: bool,
) -> RowResult:
    raw = _json_safe_raw(raw)
    try:
        student_name = _clean_str(raw.get("student_name"))
        if not student_name:
            raise ValueError("student_name is required")
        email = _clean_str(raw.get("email"))
        if not email:
            raise ValueError("email is required")
        phone_raw = _clean_str(raw.get("phone"))
        if not phone_raw:
            raise ValueError("phone is required")
        if normalize_phone(phone_raw) is None:
            raise ValueError(
                f"phone {phone_raw!r} isn't a valid number. Use a local UAE number "
                "(050 123 4567) or full international format with the country code "
                "(+20 10 1234 5678). Formatting the column as Text in Excel avoids "
                "digits being mangled."
            )
        city = _clean_str(raw.get("city"))
        if not city:
            raise ValueError("city is required")

        date_of_birth = _parse_date_of_birth(raw.get("date_of_birth"))
        grade = _clean_str(raw.get("grade"))
        organization_name = _clean_str(raw.get("organization"))

        parent_name = _clean_str(raw.get("parent_name"))
        parent_phone = _clean_str(raw.get("parent_phone"))
        parent_email = _clean_str(raw.get("parent_email"))
    except ValueError as exc:
        return RowResult(row_number=row_number, disposition="error", data=raw, reason=str(exc))

    student, evaluation = await resolve_or_create_contact(
        db, full_name=student_name, phone=phone_raw, email=email, contact_roles=["student"], city=city,
        role_event_source="import", date_of_birth=date_of_birth, grade=grade, organization_name=organization_name,
    )
    if set_contact_organization and cohort.organization_id and student.organization_id is None:
        student.organization_id = cohort.organization_id

    # Parent info is always optional here — no minor detection or enforcement
    # happens in this system at all (see MASTER_EXECUTION_PLAN.md). If given,
    # the parent becomes the guardian/payer; if not, the student pays for
    # themselves.
    payer_contact_id = None
    if parent_name and parent_phone:
        guardian, _ = await resolve_or_create_contact(
            db, full_name=parent_name, phone=parent_phone, email=parent_email, contact_roles=["parent_guardian"],
            role_event_source="import",
        )
        payer_contact_id = guardian.id
        await _ensure_guardian_relationship(db, student_id=student.id, guardian_id=guardian.id)

    disposition = "create" if evaluation.outcome == "NEW" else ("review" if evaluation.outcome == "REVIEW" else "link")

    existing_reg = (await db.execute(
        select(Registration).where(Registration.contact_id == student.id, Registration.cohort_id == cohort.id)
    )).scalars().first()
    if existing_reg is not None:
        return RowResult(row_number=row_number, disposition="already_registered", data=raw, contact_id=str(student.id))

    try:
        # register() already has its own internal SAVEPOINT around the one
        # case that needs it (the duplicate-registration IntegrityError) and
        # leaves the session usable either way — no need to wrap it again here.
        await register(
            db, contact_id=student.id, cohort_id=cohort.id, payer_contact_id=payer_contact_id,
            registered_via="import", payment_status=payment_status,
        )
    except HTTPException as exc:
        return RowResult(row_number=row_number, disposition="error", data=raw, reason=exc.detail)

    # Always the real contact id, not just for "link" rows — at commit time
    # this is what send_import_batch_emails uses to know who to email; a
    # "create" row's id is just as real as a "link" row's by the time this
    # function returns (student.id is set the moment the Contact object is
    # constructed, whether newly inserted or fetched). Only during the
    # rolled-back dry-run pass is a "create" row's id throwaway — which is
    # fine, nothing reads a dry-run preview's contact_id for anything.
    return RowResult(row_number=row_number, disposition=disposition, data=raw, contact_id=str(student.id))


def _default_payment_status(source: str) -> str:
    if source == "b2b_sheet":
        return "paid"
    # backfill: historical records, no active collection intended — 'waived'
    # is more honest than 'unpaid', which would otherwise surface as a
    # collection task for a session that already happened.
    return "waived"


def _summarize(rows: list[RowResult]) -> dict:
    summary = {"total": len(rows), "create": 0, "link": 0, "already_registered": 0, "review": 0, "error": 0}
    for r in rows:
        summary[r.disposition] += 1
    return summary


async def dry_run(
    db: AsyncSession, *, file_bytes: bytes, cohort_id: UUID, uploaded_by: UUID, source: Literal["b2b_sheet", "backfill"],
    payment_status: str | None = None, set_contact_organization: bool = False, send_emails: bool = False,
    create_lms_accounts: bool = True,
    filename: str = "upload.xlsx",
) -> ImportBatch:
    cohort = await db.get(Cohort, cohort_id)
    if cohort is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Cohort not found")

    raw_rows = parse_workbook(file_bytes)
    effective_payment_status = payment_status or _default_payment_status(source)

    results: list[RowResult] = []
    try:
        async with db.begin_nested():
            for entry in raw_rows:
                result = await _process_row(
                    db, entry["row_number"], entry["raw"], cohort, effective_payment_status, set_contact_organization,
                )
                results.append(result)
            raise _DryRunRollback()
    except _DryRunRollback:
        pass

    batch = ImportBatch(
        id=uuid4(), uploaded_by=uploaded_by, source=source, organization_id=cohort.organization_id,
        cohort_id=cohort_id, filename=filename, status="dry_run",
        counts={"summary": _summarize(results), "rows": [asdict(r) for r in results],
                "options": {"payment_status": effective_payment_status, "set_contact_organization": set_contact_organization,
                            "send_emails": send_emails, "create_lms_accounts": create_lms_accounts}},
    )
    db.add(batch)
    await db.flush()
    return batch


async def commit_batch(db: AsyncSession, batch_id: UUID) -> ImportBatch:
    batch = await db.get(ImportBatch, batch_id)
    if batch is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Import batch not found")
    if batch.status != "dry_run":
        raise HTTPException(status.HTTP_409_CONFLICT, detail=f"Batch is already {batch.status}, not dry_run")

    cohort = await db.get(Cohort, batch.cohort_id)
    options = batch.counts.get("options", {})
    payment_status = options.get("payment_status") or _default_payment_status(batch.source)
    set_contact_organization = bool(options.get("set_contact_organization"))

    results: list[RowResult] = []
    for stored in batch.counts.get("rows", []):
        if stored["disposition"] == "error":
            results.append(RowResult(**stored))
            continue
        result = await _process_row(
            db, stored["row_number"], stored["data"], cohort, payment_status, set_contact_organization,
        )
        results.append(result)

    batch.status = "committed"
    batch.counts = {**batch.counts, "summary": _summarize(results), "rows": [asdict(r) for r in results]}
    await db.flush()
    return batch


def generate_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Students"
    sheet.append([name for name, _ in TEMPLATE_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
