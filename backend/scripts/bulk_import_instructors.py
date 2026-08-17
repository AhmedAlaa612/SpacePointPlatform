"""Bulk-create instructor accounts from the boss's "Instructors" sheet
instead of making them sign up themselves (operator ask, 2026-08-17).

────────────────────────────────────────────────────────────────────────────
USAGE
────────────────────────────────────────────────────────────────────────────
    cd backend
    .venv\\Scripts\\python.exe scripts\\bulk_import_instructors.py --dry-run
    .venv\\Scripts\\python.exe scripts\\bulk_import_instructors.py

Reads whichever database `DATABASE_URL` in the active `.env` points at (same
as running the app itself) — point the dev `.env` at spacepoint_dev to test,
the prod env to run for real. There is no separate --db flag on purpose: this
script is meant to be run the same way the app runs, not pointed somewhere
unusual by accident.

--file      : path to the .xlsx (default: the sheet the operator provided).
--sheet     : worksheet name (default: "Instructors").
--dry-run   : parses, matches against the DB, and prints the full report —
              rolls back at the end instead of committing. Always run this
              first.

────────────────────────────────────────────────────────────────────────────
WHAT THIS DOES, PER ROW
────────────────────────────────────────────────────────────────────────────
1. Row has no name, or no email -> reported under "cannot import", skipped.
   (The sheet has two shapes of row mixed together: ~15 rows with a full
   name + details, and ~20 fragment rows that are just a bare email and/or
   a LinkedIn link with no name at all — those fragments cannot become an
   account without at least a name.)
2. Email already used by an existing account (case-insensitive) -> reported
   under "already exists", skipped — never touched, never re-invited.
3. Email repeated within the sheet itself -> only the first occurrence is
   considered; the rest are reported under "duplicate in sheet".
4. Otherwise: creates User(roles=[instructor] [+intern if the sheet's
   "Intern?" column says yes], must_change_password=True, a random
   password nobody is ever told), an InstructorProfile with
   instructor_since = today (the sheet has no join-date column for
   instructors — operator confirmed 2026-08-17: today is correct for
   whoever isn't already an account), a spine Contact link, and a
   card_number. Sends a welcome email with a 24h set-password link — the
   instructor is expected to log in, set their password, and fill in city
   + delivery cities themselves (operator's explicit plan, not a gap here).

Fields the sheet has that NOTHING in the current schema captures — "Major"
and "School/Uni instructor" (delivery context) — are intentionally dropped,
not silently invented a column for. Reported once in the summary so nobody
assumes they were saved somewhere. "Personal Picture" (a local filename or
Drive link, not actual image bytes this script has access to) is dropped
for the same reason — instructors can upload their own photo from Profile.

City is matched against the `cities` table (country='AE') by
case-insensitive name; on no match the raw sheet text is kept in
`city_other` instead of being lost, and the row is flagged so an admin can
add that city if it's missing rather than the person's location silently
vanishing.
"""

from __future__ import annotations

import argparse
import asyncio
import secrets
import sys
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from sqlalchemy import func, select

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.core.config import settings
from app.core.security import create_password_set_token, get_password_hash
from app.models.enums import UserRole
from app.models.inventory.city import City
from app.models.instructors.instructor_profile import InstructorProfile
from app.models.user import User
from app.services.documents.id_card import ensure_card_number
from app.services.email import send_instructor_welcome_email
from app.services.spine.identity import ensure_user_contact

DEFAULT_FILE = r"C:\Users\ahmed\Downloads\Untitled spreadsheet (4).xlsx"
DEFAULT_SHEET = "Instructors"

# Sheet uses abbreviations the `cities` table doesn't necessarily store —
# extend this if a dry-run reports more unmatched cities than expected.
_CITY_ALIASES = {"rak": "ras al khaimah", "uaq": "umm al quwain"}


@dataclass
class Row:
    excel_row: int
    legacy_id: str | None
    name: str | None
    major: str | None
    city_raw: str | None
    email: str | None
    phone_raw: object
    linkedin: str | None
    is_intern: bool


@dataclass
class Report:
    cannot_import: list[str] = field(default_factory=list)
    duplicate_in_sheet: list[str] = field(default_factory=list)
    already_exists: list[str] = field(default_factory=list)
    city_unmatched: list[str] = field(default_factory=list)
    created: list[str] = field(default_factory=list)
    emails_sent: int = 0
    emails_failed: list[str] = field(default_factory=list)

    def print_summary(self, *, dry_run: bool) -> None:
        print("\n" + "=" * 78)
        print(f"BULK INSTRUCTOR IMPORT {'(DRY RUN — nothing was committed)' if dry_run else '(LIVE RUN)'}")
        print("=" * 78)
        print(f"To create: {len(self.created)}")
        for s in self.created:
            print(f"  + {s}")
        print(f"\nAlready exist (skipped, {len(self.already_exists)}):")
        for s in self.already_exists:
            print(f"  = {s}")
        print(f"\nDuplicate email within the sheet ({len(self.duplicate_in_sheet)}):")
        for s in self.duplicate_in_sheet:
            print(f"  ~ {s}")
        print(f"\nCannot import — missing name and/or email ({len(self.cannot_import)}):")
        for s in self.cannot_import:
            print(f"  ! {s}")
        if self.city_unmatched:
            print(f"\nCity text didn't match the `cities` table, kept as free text ({len(self.city_unmatched)}):")
            for s in self.city_unmatched:
                print(f"  ? {s}")
        if not dry_run:
            print(f"\nWelcome emails sent: {self.emails_sent}")
            if self.emails_failed:
                print(f"Welcome emails FAILED to send ({len(self.emails_failed)}) — account was still created:")
                for s in self.emails_failed:
                    print(f"  x {s}")
        print("\nDropped fields not captured anywhere in the schema: Major, School/Uni")
        print("instructor type, Personal Picture. Not saved, not invented a column for.")
        print("=" * 78 + "\n")


def _clean_email(raw: object) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if s.lower().startswith("mailto:"):
        s = s[len("mailto:"):].strip()
    return s or None


def _clean_phone(raw: object) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, float) and raw.is_integer():
        s = str(int(raw))
    else:
        s = str(raw).strip()
    if not s or s.upper() == "N/A":
        return None
    return s if s.startswith("+") else f"+{s}"


def _clean_linkedin(raw: object) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    return s if s.lower().startswith("http") else None


def _parse_rows(path: str, sheet_name: str) -> list[Row]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found. Sheets in file: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows: list[Row] = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        legacy_id, name, major, city_raw, email_raw, phone_raw, linkedin_raw = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
        intern_raw = r[8] if len(r) > 8 else None
        if all(v is None for v in r):
            continue
        rows.append(Row(
            excel_row=i,
            legacy_id=(str(int(legacy_id)) if isinstance(legacy_id, float) else (str(legacy_id) if legacy_id else None)),
            name=(str(name).strip() if name else None) or None,
            major=(str(major).strip() if major else None) or None,
            city_raw=(str(city_raw).strip() if city_raw else None) or None,
            email=_clean_email(email_raw),
            phone_raw=phone_raw,
            linkedin=_clean_linkedin(linkedin_raw),
            is_intern=str(intern_raw).strip().lower() in ("yes", "y", "true") if intern_raw else False,
        ))
    return rows


async def _resolve_city(db, city_raw: str | None) -> tuple[uuid.UUID | None, str | None]:
    """Returns (city_id, city_other) — exactly one populated, matching the
    same mutual-exclusivity convention as User.city_id/city_other elsewhere."""
    if not city_raw:
        return None, None
    cities = (await db.execute(select(City).where(City.country == "AE"))).scalars().all()
    by_name = {c.name.strip().lower(): c.id for c in cities}
    key = city_raw.strip().lower()
    key = _CITY_ALIASES.get(key, key)
    if key in by_name:
        return by_name[key], None
    return None, city_raw


async def run(args) -> Report:
    from app.db.session import AsyncSessionLocal

    report = Report()
    rows = _parse_rows(args.file, args.sheet)

    seen_emails: dict[str, Row] = {}
    usable: list[Row] = []
    for row in rows:
        label = f"row {row.excel_row}: {row.name or '(no name)'} / {row.email or '(no email)'}"
        if not row.name or not row.email:
            report.cannot_import.append(label)
            continue
        key = row.email.lower()
        if key in seen_emails:
            report.duplicate_in_sheet.append(f"{label} — first seen at row {seen_emails[key].excel_row}")
            continue
        seen_emails[key] = row
        usable.append(row)

    async with AsyncSessionLocal() as db:
        for row in usable:
            existing = (await db.execute(
                select(User.id, User.full_name).where(func.lower(User.email) == row.email.lower())
            )).first()
            if existing:
                report.already_exists.append(f"{row.name} <{row.email}> — already an account ({existing.full_name}, id={existing.id})")
                continue

            city_id, city_other = await _resolve_city(db, row.city_raw)
            if row.city_raw and city_id is None:
                report.city_unmatched.append(f"{row.name}: '{row.city_raw}' — kept as free text")

            roles = [UserRole.instructor] + ([UserRole.intern] if row.is_intern else [])
            today = datetime.now(timezone.utc).date()

            user = User(
                full_name=row.name,
                email=row.email,
                password_hash=get_password_hash(secrets.token_urlsafe(24)),
                roles=roles,
                status="active",
                must_change_password=True,
                phone=_clean_phone(row.phone_raw),
                country="AE",
                city_id=city_id,
                city_other=city_other,
                linkedin_url=row.linkedin,
            )
            db.add(user)
            await db.flush()
            await ensure_user_contact(db, user, source="bulk_import_instructor")
            await ensure_card_number(db, user)
            db.add(InstructorProfile(user_id=user.id, instructor_since=today))

            report.created.append(
                f"{row.name} <{row.email}> (legacy id {row.legacy_id or '?'}, card #{user.card_number}, roles={[r.value for r in roles]})"
            )

            if not args.dry_run:
                token = create_password_set_token(user.id)
                link = f"{settings.FRONTEND_URL}/set-password?token={token}"
                sent = await send_instructor_welcome_email(user.email, user.full_name, link)
                if sent:
                    report.emails_sent += 1
                else:
                    report.emails_failed.append(f"{row.name} <{row.email}>")

        if args.dry_run:
            await db.rollback()
        else:
            await db.commit()

    return report


def main() -> None:
    p = argparse.ArgumentParser(description="Bulk-create instructor accounts from the boss's Instructors sheet")
    p.add_argument("--file", default=DEFAULT_FILE)
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    report = asyncio.run(run(args))
    report.print_summary(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
