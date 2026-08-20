"""Bulk-create intern accounts from the "Interns" sheet of the operator's
instructors/interns workbook, mirroring bulk_import_instructors.py (operator
ask, 2026-08-20 — same reasoning: don't make ~76 people sign up themselves).

────────────────────────────────────────────────────────────────────────────
USAGE
────────────────────────────────────────────────────────────────────────────
    cd backend
    .venv\\Scripts\\python.exe scripts\\bulk_import_interns.py --dry-run
    .venv\\Scripts\\python.exe scripts\\bulk_import_interns.py

Reads whichever database `DATABASE_URL` in the active `.env` points at — same
convention as bulk_import_instructors.py, no separate --db flag on purpose.

--file      : path to the .xlsx (default: the sheet the operator provided).
--sheet     : worksheet name (default: "Interns").
--dry-run   : parses, matches against the DB, and prints the full report —
              rolls back at the end instead of committing. Always run this
              first.
--csv-out   : on a REAL run only, writes name/email/set-password-link for
              every account created to this CSV path (default:
              set_password_links.csv) — sensitive, same as the instructor
              script's file: anyone holding a link can set that account's
              password within 3 days. Delete it once everyone's confirmed in.

────────────────────────────────────────────────────────────────────────────
WHAT THIS DOES, PER ROW
────────────────────────────────────────────────────────────────────────────
1. Row has no name or no email -> "cannot import", skipped. (Unlike the
   Instructors sheet, this one has no fragment rows — every real row has
   both; only the sheet's trailing fully-blank rows hit this branch.)
2. Contract Status is neither "Active" nor "Signed" -> "not onboarded",
   skipped (operator call, 2026-08-20: onboard both statuses today — all 76
   real rows are one or the other, this is just future-proofing a re-run
   against an updated sheet).
3. Email already used by an existing account (case-insensitive, checked
   against the LIVE DB — the sheet's own "Registered" column is NOT
   trusted, it's informational only) -> "already exists", skipped.
4. Email repeated within the sheet -> only the first occurrence is
   considered; the rest are "duplicate in sheet".
5. Otherwise: creates User(roles=[intern], must_change_password=True, a
   random password nobody is ever told), an InternProfile with ref_number/
   university_id_number/department/start_date from the sheet — data only,
   no internship letter or signature (operator call: these people already
   started, some back in 2024; a fresh letter now would be retroactive and
   odd), a spine Contact link, and a card_number. Sends a welcome email with
   a 3-day set-password link.

Ref Number: the sheet's earliest ~8 rows carry a placeholder DATE in this
column instead of a real "N/YYYY" reference (a pre-existing data artifact,
not something this script introduces) — stored as NULL rather than a
garbled date-as-ref-number, and reported once in the summary so it's not
mistaken for "no ref number was ever assigned". Everything from a normal
"N/YYYY" string is kept as-is; storing it verbatim (not renumbering) is what
lets services/internship/ref_number.py seed its per-year counter from these
imported rows and continue after the highest real one instead of colliding.

Student ID -> `university_id_number`: kept as free text (mixes plain
numbers, "N/A" -> dropped to NULL, and alphanumeric university IDs like
"2024A7PS0271U") — never coerced to a number.

Department: free text, no controlled list on the source sheet; NULL when
blank (~30% of rows).
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import os
import re
import secrets
import sys
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from sqlalchemy import func, select

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.core.config import settings
from app.core.security import create_password_set_token, get_password_hash
from app.models.enums import UserRole
from app.models.internship import InternProfile
from app.models.user import User
from app.services.documents.id_card import ensure_card_number
from app.services.email import send_intern_welcome_email
from app.services.spine.identity import ensure_user_contact

DEFAULT_FILE = r"C:\Users\ahmed\Downloads\Untitled spreadsheet (4).xlsx"
DEFAULT_SHEET = "Interns"

_ONBOARD_STATUSES = {"active", "signed"}
_REF_NUMBER_RE = re.compile(r"^\d+/\d{4}$")
_ORDINAL_SUFFIX_RE = re.compile(r"(\d+)(st|nd|rd|th)", re.IGNORECASE)


@dataclass
class Row:
    excel_row: int
    name: str | None
    ref_number_raw: object
    joining_date_raw: object
    email: str | None
    student_id_raw: object
    department: str | None
    contract_status: str | None
    registered_raw: object


@dataclass
class Report:
    cannot_import: list[str] = field(default_factory=list)
    not_onboarded_status: list[str] = field(default_factory=list)
    duplicate_in_sheet: list[str] = field(default_factory=list)
    already_exists: list[str] = field(default_factory=list)
    ref_number_missing: list[str] = field(default_factory=list)
    created: list[str] = field(default_factory=list)
    emails_sent: int = 0
    emails_failed: list[str] = field(default_factory=list)
    csv_path: str | None = None

    def print_summary(self, *, dry_run: bool) -> None:
        print("\n" + "=" * 78)
        print(f"BULK INTERN IMPORT {'(DRY RUN — nothing was committed)' if dry_run else '(LIVE RUN)'}")
        print("=" * 78)
        print(f"To create: {len(self.created)}")
        for s in self.created:
            print(f"  + {s}")
        print(f"\nAlready exist (skipped, {len(self.already_exists)}):")
        for s in self.already_exists:
            print(f"  = {s}")
        print(f"\nDuplicate email within the sheet ({len(self.duplicate_in_sheet)}):")
        for s in self.duplicate_in_sheet:
            print(f"  ~ {s}")
        print(f"\nContract Status not Active/Signed, not onboarded ({len(self.not_onboarded_status)}):")
        for s in self.not_onboarded_status:
            print(f"  ! {s}")
        print(f"\nCannot import — missing name and/or email ({len(self.cannot_import)}):")
        for s in self.cannot_import:
            print(f"  ! {s}")
        if self.ref_number_missing:
            print(f"\nNo real ref number on file — placeholder date in sheet, stored as NULL ({len(self.ref_number_missing)}):")
            for s in self.ref_number_missing:
                print(f"  ? {s}")
        if not dry_run:
            print(f"\nWelcome emails sent: {self.emails_sent}")
            if self.emails_failed:
                print(f"Welcome emails FAILED to send ({len(self.emails_failed)}) — account was still created:")
                for s in self.emails_failed:
                    print(f"  x {s}")
            if self.csv_path:
                print(f"\nSet-password links written to: {self.csv_path}")
                print("This file is as sensitive as a password list (anyone holding a link can set")
                print("that account's password within 3 days) — delete it once everyone's confirmed in.")
        print("=" * 78 + "\n")


def _blank(v: object) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _clean_email(raw: object) -> str | None:
    if _blank(raw):
        return None
    return str(raw).strip()


def _clean_student_id(raw: object) -> str | None:
    if _blank(raw):
        return None
    s = str(raw).strip()
    if s.upper() == "N/A":
        return None
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return s


def _clean_ref_number(raw: object) -> tuple[str | None, bool]:
    """Returns (ref_number, was_missing). A datetime here is a pre-existing
    sheet artifact (placeholder date instead of a real "N/YYYY" ref) — kept
    as NULL, not stringified, so it can never be mistaken for a real
    reference or collide with the auto-increment counter's seeding scan."""
    if isinstance(raw, datetime):
        return None, True
    if _blank(raw):
        return None, True
    s = str(raw).strip()
    return s, not bool(_REF_NUMBER_RE.match(s))


def _parse_joining_date(raw: object) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if _blank(raw):
        return None
    s = _ORDINAL_SUFFIX_RE.sub(r"\1", str(raw).strip())
    for fmt in ("%B %d, %Y", "%d %B %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_rows(path: str, sheet_name: str) -> list[Row]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found. Sheets in file: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows: list[Row] = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in r):
            continue
        name, ref_number_raw, joining_date_raw, email_raw, student_id_raw, department_raw, status_raw, registered_raw = (
            r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7] if len(r) > 7 else None,
        )
        rows.append(Row(
            excel_row=i,
            name=(str(name).strip() if name else None) or None,
            ref_number_raw=ref_number_raw,
            joining_date_raw=joining_date_raw,
            email=_clean_email(email_raw),
            student_id_raw=student_id_raw,
            department=(str(department_raw).strip() if department_raw else None) or None,
            contract_status=(str(status_raw).strip() if status_raw else None) or None,
            registered_raw=registered_raw,
        ))
    return rows


async def run(args) -> Report:
    from app.db.session import AsyncSessionLocal

    report = Report()
    rows = _parse_rows(args.file, args.sheet)

    seen_emails: dict[str, Row] = {}
    usable: list[Row] = []
    for row in rows:
        label = f"row {row.excel_row}: {row.name or '(no name)'} / {row.email or '(no email)'}"
        if not row.name or not row.email:
            report.cannot_import.append(label)
            continue
        if (row.contract_status or "").strip().lower() not in _ONBOARD_STATUSES:
            report.not_onboarded_status.append(f"{label} — status={row.contract_status!r}")
            continue
        key = row.email.lower()
        if key in seen_emails:
            report.duplicate_in_sheet.append(f"{label} — first seen at row {seen_emails[key].excel_row}")
            continue
        seen_emails[key] = row
        usable.append(row)

    set_password_links: list[tuple[str, str, str]] = []  # (name, email, link) — real run only

    async with AsyncSessionLocal() as db:
        for row in usable:
            existing = (await db.execute(
                select(User.id, User.full_name).where(func.lower(User.email) == row.email.lower())
            )).first()
            if existing:
                report.already_exists.append(f"{row.name} <{row.email}> — already an account ({existing.full_name}, id={existing.id})")
                continue

            ref_number, ref_missing = _clean_ref_number(row.ref_number_raw)
            if ref_missing:
                report.ref_number_missing.append(f"{row.name} <{row.email}>")

            user = User(
                full_name=row.name,
                email=row.email,
                password_hash=get_password_hash(secrets.token_urlsafe(24)),
                roles=[UserRole.intern],
                status="active",
                must_change_password=True,
            )
            db.add(user)
            await db.flush()
            await ensure_user_contact(db, user, source="bulk_import_intern")
            await ensure_card_number(db, user)
            db.add(InternProfile(
                user_id=user.id,
                ref_number=ref_number,
                university_id_number=_clean_student_id(row.student_id_raw),
                department=row.department,
                start_date=_parse_joining_date(row.joining_date_raw),
            ))

            report.created.append(
                f"{row.name} <{row.email}> (ref {ref_number or '—'}, card #{user.card_number}, status={row.contract_status})"
            )

            if not args.dry_run:
                token = create_password_set_token(user.id, expires_delta=timedelta(days=3))
                link = f"{settings.FRONTEND_URL}/set-password?token={token}"
                set_password_links.append((row.name, row.email, link))
                sent = await send_intern_welcome_email(user.email, user.full_name, link)
                if sent:
                    report.emails_sent += 1
                else:
                    report.emails_failed.append(f"{row.name} <{row.email}>")

        if args.dry_run:
            await db.rollback()
        else:
            await db.commit()

    if set_password_links:
        csv_path = Path(args.csv_out)
        with open(csv_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["full_name", "email", "set_password_link"])
            writer.writerows(set_password_links)
        try:
            os.chmod(csv_path, 0o600)  # best-effort — no-op on Windows
        except OSError:
            pass
        report.csv_path = str(csv_path)

    return report


def main() -> None:
    p = argparse.ArgumentParser(description="Bulk-create intern accounts from the operator's Interns sheet")
    p.add_argument("--file", default=DEFAULT_FILE)
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--csv-out", default="set_password_links_interns.csv")
    args = p.parse_args()

    report = asyncio.run(run(args))
    report.print_summary(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
