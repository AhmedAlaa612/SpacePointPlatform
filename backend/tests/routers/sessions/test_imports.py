"""Router-level tests for the bulk importer endpoints (V2 R2-2's
routers/sessions/imports.py) — the service layer (services/sessions/importer.py)
already has thorough tests in tests/services/sessions/test_importer.py; these
cover the actual HTTP surface (multipart upload, dry-run/commit round trip via
real requests, the cohort_id list filter) that only a real client can exercise.
"""

import io
import uuid

import httpx
import openpyxl
import pytest
from sqlalchemy import select

from app.db.session import get_db
from app.main import app
from app.models.sessions.cohort import Cohort
from app.models.sessions.program import Program
from app.models.sessions.registration import Registration
from app.services.sessions.importer import TEMPLATE_COLUMNS
from app.workers.settings import get_arq_redis


# `client` (Redis-free) and `arq_client` (real ARQ pool) live in
# tests/conftest.py. The local copy that used to be here bound *every* test in
# this file to a live Redis, including ones that never enqueue anything (I0-1b).


async def _make_cohort(db) -> Cohort:
    program = Program(
        id=uuid.uuid4(), code=f"IMP-RT-{uuid.uuid4().hex[:8]}", name="Import Router Test Program",
        program_type="workshop", pricing_model="free", active=True,
    )
    db.add(program)
    await db.flush()
    cohort = Cohort(
        id=uuid.uuid4(), program_id=program.id, name="Import Router Test Cohort",
        status="registration_open", visibility="public",
    )
    db.add(cohort)
    await db.flush()
    return cohort


def _build_workbook(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append([name for name, _ in TEMPLATE_COLUMNS])
    for row in rows:
        sheet.append([row.get(name) for name, _ in TEMPLATE_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _student_row(**overrides) -> dict:
    row = dict(
        student_name="Router Test Student", email="router.student@example.com",
        phone="050 222 3333", city="Dubai",
        parent_name=None, parent_phone=None, parent_email=None,
    )
    row.update(overrides)
    return row


@pytest.mark.asyncio
async def test_download_template_requires_operations_role(client, other_role_headers):
    resp = await client.get("/sessions/imports/template", headers=other_role_headers)
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_download_template_returns_an_xlsx(client, operations_headers):
    resp = await client.get("/sessions/imports/template", headers=operations_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
    # A real workbook — openpyxl can load it back without error.
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    header = [c.value for c in next(wb.active.iter_rows(max_row=1))]
    assert header == [name for name, _ in TEMPLATE_COLUMNS]


@pytest.mark.asyncio
async def test_dry_run_then_commit_via_real_multipart_upload(db, client, operations_headers):
    cohort = await _make_cohort(db)
    file_bytes = _build_workbook([_student_row()])

    dry_run_resp = await client.post(
        "/sessions/imports/dry-run",
        data={"cohort_id": str(cohort.id), "source": "b2b_sheet", "send_emails": "false"},
        files={"file": ("students.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=operations_headers,
    )
    assert dry_run_resp.status_code == 200, dry_run_resp.text
    batch = dry_run_resp.json()
    assert batch["status"] == "dry_run"
    assert len(batch["rows"]) == 1
    assert batch["rows"][0]["disposition"] == "create"

    # Nothing committed yet.
    count = (await db.execute(select(Registration).where(Registration.cohort_id == cohort.id))).scalars().all()
    assert count == []

    commit_resp = await client.post(f"/sessions/imports/{batch['id']}/commit", headers=operations_headers)
    assert commit_resp.status_code == 200, commit_resp.text
    committed = commit_resp.json()
    assert committed["status"] == "committed"

    registrations = (await db.execute(select(Registration).where(Registration.cohort_id == cohort.id))).scalars().all()
    assert len(registrations) == 1


@pytest.mark.asyncio
async def test_list_batches_filters_by_cohort_id(db, client, operations_headers):
    cohort_a = await _make_cohort(db)
    cohort_b = await _make_cohort(db)

    for cohort in (cohort_a, cohort_b):
        file_bytes = _build_workbook([_student_row(email=f"{uuid.uuid4().hex}@example.com")])
        resp = await client.post(
            "/sessions/imports/dry-run",
            data={"cohort_id": str(cohort.id), "source": "b2b_sheet", "send_emails": "false"},
            files={"file": ("students.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=operations_headers,
        )
        assert resp.status_code == 200, resp.text

    resp = await client.get("/sessions/imports", params={"cohort_id": str(cohort_a.id)}, headers=operations_headers)
    assert resp.status_code == 200
    batches = resp.json()
    assert len(batches) == 1
    assert batches[0]["cohort_id"] == str(cohort_a.id)
